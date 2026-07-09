"""
reports/generator.py
--------------------
Generates styled downloadable Excel reports:
  - Pending Maintenance List  (by month or year)
  - Balance Sheet             (by month or year)

Both functions return raw bytes so Streamlit can serve them
directly via st.download_button without writing to disk.
"""

import io
import sys
import pandas as pd
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))
from server.data_loader import (
    get_pending_maintenance,
    get_collected_maintenance,
    get_balance_sheet,
    get_expense_summary,
    load_maintenance,
)

# ── Colour palette ─────────────────────────────────────────────────────────────
C_HEADER_BG   = "1E3A5F"   # dark navy  — header row background
C_HEADER_FG   = "FFFFFF"   # white      — header row text
C_TITLE_BG    = "2D6A9F"   # mid blue   — title / section header
C_TITLE_FG    = "FFFFFF"
C_SUBHEAD_BG  = "D6E4F0"   # light blue — sub-section header
C_SUBHEAD_FG  = "1E3A5F"
C_ALT_ROW     = "F0F6FB"   # very light blue — alternating row
C_UNPAID_BG   = "FFF3CD"   # pale amber — unpaid / pending row
C_UNPAID_FG   = "856404"
C_PAID_FG     = "155724"   # dark green — paid status
C_TOTAL_BG    = "E8F4EA"   # light green — total row
C_SURPLUS_FG  = "155724"   # green  — positive balance
C_DEFICIT_FG  = "842029"   # red    — negative balance
C_BORDER      = "BDD7EE"   # light border colour

# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color="000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_thin() -> Border:
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width

def _title_row(ws, text: str, ncols: int, row: int):
    """Write a merged title row across ncols columns."""
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill      = _fill(C_TITLE_BG)
    cell.font      = _font(bold=True, size=14, color=C_TITLE_FG)
    cell.alignment = _align(h="center")
    ws.row_dimensions[row].height = 28

def _section_header(ws, text: str, ncols: int, row: int):
    """Write a merged section-header row."""
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill      = _fill(C_SUBHEAD_BG)
    cell.font      = _font(bold=True, size=11, color=C_SUBHEAD_FG)
    cell.alignment = _align(h="left")
    ws.row_dimensions[row].height = 20

def _header_row(ws, headers: list[str], row: int):
    """Write a styled column-header row."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill      = _fill(C_HEADER_BG)
        cell.font      = _font(bold=True, color=C_HEADER_FG)
        cell.alignment = _align(h="center")
        cell.border    = _border_thin()
    ws.row_dimensions[row].height = 18

def _data_cell(ws, row: int, col: int, value,
               bold=False, h_align="left", bg=None,
               fg="000000", number_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, color=fg)
    cell.alignment = _align(h=h_align)
    cell.border    = _border_thin()
    if bg:
        cell.fill = _fill(bg)
    if number_fmt:
        cell.number_format = number_fmt
    return cell

def _generated_note(ws, row: int, ncols: int):
    """Add a small 'Generated on' footer row."""
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1,
                   value=f"Generated on: {datetime.now().strftime('%d %b %Y, %I:%M %p')}")
    cell.font      = _font(size=9, italic=True, color="57606A")
    cell.alignment = _align(h="right")


# ══════════════════════════════════════════════════════════════════════════════
#  Report 1 — Pending Maintenance
# ══════════════════════════════════════════════════════════════════════════════

def generate_pending_excel(month: str = None, year: int = None) -> bytes:
    """
    Generate a styled Excel report of pending (unpaid) maintenance.
    Returns file bytes for Streamlit download.
    """
    period = month or (str(year) if year else "All")

    # ── Fetch data ─────────────────────────────────────────────
    pending_df   = get_pending_maintenance(month=month, year=year)
    collected_df = get_collected_maintenance(month=month, year=year)

    # Count total occupied flats for the period
    maint_df = load_maintenance()
    maint_df = maint_df[~maint_df["is_vacant"]]
    if month:
        from server.data_loader import _normalise_month
        month = _normalise_month(month)
        maint_df = maint_df[maint_df["month"].str.lower() == month.lower()]
    elif year:
        maint_df = maint_df[maint_df["year"] == int(year)]
    total_flats   = len(maint_df)
    paid_count    = len(collected_df)
    pending_count = len(pending_df)

    # ── Workbook setup ─────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pending Maintenance"
    ws.sheet_view.showGridLines = False

    # Year query → grouped shape (flat_number, name, months_pending, months_count, total_due)
    # Month query → per-row shape (month, flat_number, name, amount_due, late_charges)
    is_year_query = year and not month

    if is_year_query:
        NCOLS = 5
        _set_col_width(ws, 1, 12)   # Flat No
        _set_col_width(ws, 2, 30)   # Resident Name
        _set_col_width(ws, 3, 10)   # Months Unpaid
        _set_col_width(ws, 4, 45)   # Months (list)
        _set_col_width(ws, 5, 16)   # Total Due
    else:
        NCOLS = 5
        _set_col_width(ws, 1, 12)   # Flat No
        _set_col_width(ws, 2, 30)   # Resident Name
        _set_col_width(ws, 3, 16)   # Month
        _set_col_width(ws, 4, 16)   # Amount Due
        _set_col_width(ws, 5, 16)   # Late Charges

    r = 1
    # Title
    _title_row(ws, f"PENDING MAINTENANCE REPORT — {period.upper()}", NCOLS, r); r += 1

    # Summary strip
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    summary_text = (
        f"Total Flats: {total_flats}   |   "
        f"Paid: {paid_count}   |   "
        f"Pending: {pending_count}"
    )
    sc = ws.cell(row=r, column=1, value=summary_text)
    sc.fill      = _fill(C_SUBHEAD_BG)
    sc.font      = _font(size=10, color=C_SUBHEAD_FG)
    sc.alignment = _align(h="center")
    ws.row_dimensions[r].height = 18
    r += 1

    r += 1  # blank row

    if pending_df.empty:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        cell = ws.cell(row=r, column=1,
                       value=f"✅ All flats have paid maintenance for {period}!")
        cell.font      = _font(bold=True, size=12, color=C_PAID_FG)
        cell.alignment = _align(h="center")
        r += 2
    elif is_year_query:
        # ── Year-grouped unpaid section ────────────────────────
        _section_header(ws, "  UNPAID FLATS (YEAR SUMMARY)", NCOLS, r); r += 1
        headers = ["Flat No.", "Resident Name", "Months Unpaid",
                   "Unpaid Months", "Total Due (Rs.)"]
        _header_row(ws, headers, r); r += 1

        for i, row_data in pending_df.iterrows():
            bg = C_UNPAID_BG if i % 2 == 0 else None
            _data_cell(ws, r, 1, int(row_data["flat_number"]),
                       h_align="center", bg=bg, bold=True)
            _data_cell(ws, r, 2, str(row_data["name"]), bg=bg)
            _data_cell(ws, r, 3, int(row_data["months_count"]),
                       h_align="center", bg=bg, bold=True)
            _data_cell(ws, r, 4, str(row_data["months_pending"]),
                       bg=bg)
            _data_cell(ws, r, 5, float(row_data["total_due"]),
                       h_align="right", bg=bg, number_fmt='#,##0.00')
            r += 1

        # Total row
        grand_total = float(pending_df["total_due"].sum())
        _data_cell(ws, r, 1, "TOTAL", bold=True, h_align="center", bg=C_TOTAL_BG)
        _data_cell(ws, r, 2, f"{pending_count} flat(s) pending",
                   bold=True, bg=C_TOTAL_BG)
        ws.cell(row=r, column=3).fill = _fill(C_TOTAL_BG)
        ws.cell(row=r, column=4).fill = _fill(C_TOTAL_BG)
        _data_cell(ws, r, 5, grand_total, bold=True, h_align="right",
                   bg=C_TOTAL_BG, number_fmt='#,##0.00')
        r += 2
    else:
        # ── Month per-row unpaid section ───────────────────────
        _section_header(ws, "  UNPAID FLATS", NCOLS, r); r += 1
        headers = ["Flat No.", "Resident Name", "Month",
                   "Amount Due (Rs.)", "Late Charges (Rs.)"]
        _header_row(ws, headers, r); r += 1

        for i, row_data in pending_df.iterrows():
            bg = C_UNPAID_BG if i % 2 == 0 else None
            _data_cell(ws, r, 1, int(row_data["flat_number"]),
                       h_align="center", bg=bg, bold=True)
            _data_cell(ws, r, 2, str(row_data["name"]), bg=bg)
            _data_cell(ws, r, 3, str(row_data["month"]),
                       h_align="center", bg=bg)
            _data_cell(ws, r, 4, float(row_data["amount_due"]),
                       h_align="right", bg=bg, number_fmt='#,##0.00')
            _data_cell(ws, r, 5, float(row_data["late_charges"]),
                       h_align="right", bg=bg, number_fmt='#,##0.00')
            r += 1

        # Total row
        total_due = float(pending_df["amount_due"].sum())
        total_lc  = float(pending_df["late_charges"].sum())
        _data_cell(ws, r, 1, "TOTAL", bold=True, h_align="center", bg=C_TOTAL_BG)
        _data_cell(ws, r, 2, f"{pending_count} flat(s) pending",
                   bold=True, bg=C_TOTAL_BG)
        ws.cell(row=r, column=3).fill = _fill(C_TOTAL_BG)
        _data_cell(ws, r, 4, total_due, bold=True, h_align="right",
                   bg=C_TOTAL_BG, number_fmt='#,##0.00')
        _data_cell(ws, r, 5, total_lc, bold=True, h_align="right",
                   bg=C_TOTAL_BG, number_fmt='#,##0.00')
        r += 2

    # Paid flats section
    if not collected_df.empty:
        _section_header(ws, "  PAID FLATS", NCOLS, r); r += 1
        paid_headers = ["Flat No.", "Resident Name", "Month",
                        "Amount Paid (Rs.)", "Payment Date"]
        _header_row(ws, paid_headers, r); r += 1

        for i, row_data in collected_df.iterrows():
            bg = C_ALT_ROW if i % 2 == 0 else None
            _data_cell(ws, r, 1, int(row_data["flat_number"]),
                       h_align="center", bg=bg)
            _data_cell(ws, r, 2, str(row_data["name"]), bg=bg)
            _data_cell(ws, r, 3, str(row_data["month"]),
                       h_align="center", bg=bg)
            _data_cell(ws, r, 4, float(row_data["amount"]),
                       h_align="right", bg=bg, number_fmt='#,##0.00',
                       fg=C_PAID_FG)
            date_val = row_data["date"]
            date_str = date_val.strftime("%d %b %Y") if (hasattr(date_val, "strftime") and not pd.isnull(date_val)) else ""
            _data_cell(ws, r, 5, date_str, h_align="center", bg=bg)
            r += 1

        r += 1

    _generated_note(ws, r, NCOLS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  Report 2 — Balance Sheet
# ══════════════════════════════════════════════════════════════════════════════

def generate_balance_excel(month: str = None, year: int = None) -> bytes:
    """
    Generate a styled Excel balance sheet report.
    Returns file bytes for Streamlit download.
    """
    period = month or (str(year) if year else "All Time")

    # ── Fetch data ─────────────────────────────────────────────
    bs      = get_balance_sheet(month=month, year=year)
    exp_df  = get_expense_summary(month=month, year=year)

    total_collected = bs["total_collected"]
    total_expenses  = bs["total_expenses"]
    net_balance     = bs["net_balance"]
    breakdown       = bs["expense_breakdown"]

    # ── Workbook setup ─────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.sheet_view.showGridLines = False

    NCOLS = 4
    _set_col_width(ws, 1, 6)    # #
    _set_col_width(ws, 2, 35)   # Description
    _set_col_width(ws, 3, 22)   # Amount
    _set_col_width(ws, 4, 22)   # Notes

    r = 1
    _title_row(ws, f"BALANCE SHEET — {period.upper()}", NCOLS, r); r += 1
    r += 1  # blank

    # ── INCOME section ─────────────────────────────────────────
    _section_header(ws, "  INCOME", NCOLS, r); r += 1
    _header_row(ws, ["#", "Description", "Amount (Rs.)", "Notes"], r); r += 1

    # Maintenance collected
    _data_cell(ws, r, 1, "1", h_align="center", bg=C_ALT_ROW)
    _data_cell(ws, r, 2, "Maintenance Collected", bg=C_ALT_ROW)
    _data_cell(ws, r, 3, total_collected, h_align="right",
               bg=C_ALT_ROW, number_fmt='#,##0.00',
               fg=C_PAID_FG, bold=True)
    _data_cell(ws, r, 4, f"Period: {period}", bg=C_ALT_ROW,
               fg="57606A")
    r += 1

    # Income total
    _data_cell(ws, r, 1, "", bg=C_TOTAL_BG, bold=True)
    _data_cell(ws, r, 2, "TOTAL INCOME", bg=C_TOTAL_BG, bold=True)
    _data_cell(ws, r, 3, total_collected, h_align="right",
               bg=C_TOTAL_BG, bold=True,
               number_fmt='#,##0.00', fg=C_PAID_FG)
    ws.cell(row=r, column=4).fill = _fill(C_TOTAL_BG)
    r += 2

    # ── EXPENSES section ───────────────────────────────────────
    _section_header(ws, "  EXPENSES", NCOLS, r); r += 1
    _header_row(ws, ["#", "Expense Category", "Amount (Rs.)", "% of Total"], r); r += 1

    for idx, item in enumerate(breakdown, 1):
        bg = C_ALT_ROW if idx % 2 == 0 else None
        pct = (item["total"] / total_expenses * 100) if total_expenses else 0
        _data_cell(ws, r, 1, str(idx), h_align="center", bg=bg)
        _data_cell(ws, r, 2, item["expense_type"], bg=bg)
        _data_cell(ws, r, 3, item["total"], h_align="right",
                   bg=bg, number_fmt='#,##0.00')
        _data_cell(ws, r, 4, f"{pct:.1f}%", h_align="center", bg=bg,
                   fg="57606A")
        r += 1

    # Expenses total
    _data_cell(ws, r, 1, "", bg=C_TOTAL_BG, bold=True)
    _data_cell(ws, r, 2, "TOTAL EXPENSES", bg=C_TOTAL_BG, bold=True)
    _data_cell(ws, r, 3, total_expenses, h_align="right",
               bg=C_TOTAL_BG, bold=True,
               number_fmt='#,##0.00', fg=C_DEFICIT_FG)
    ws.cell(row=r, column=4).fill = _fill(C_TOTAL_BG)
    r += 2

    # ── NET BALANCE section ────────────────────────────────────
    _section_header(ws, "  NET BALANCE", NCOLS, r); r += 1

    bal_fg = C_SURPLUS_FG if net_balance >= 0 else C_DEFICIT_FG
    bal_bg = "E8F4EA"      if net_balance >= 0 else "FDECEA"
    bal_label = "SURPLUS"  if net_balance >= 0 else "DEFICIT"

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell = ws.cell(row=r, column=1,
                   value=f"NET {bal_label} (Income − Expenses)")
    cell.fill      = _fill(bal_bg)
    cell.font      = _font(bold=True, size=13, color=bal_fg)
    cell.alignment = _align(h="left")
    cell.border    = _border_thin()

    net_cell = ws.cell(row=r, column=3, value=net_balance)
    net_cell.fill         = _fill(bal_bg)
    net_cell.font         = _font(bold=True, size=13, color=bal_fg)
    net_cell.alignment    = _align(h="right")
    net_cell.number_format = '#,##0.00'
    net_cell.border        = _border_thin()

    status_cell = ws.cell(row=r, column=4,
                          value="✅ Surplus" if net_balance >= 0 else "⚠️ Deficit")
    status_cell.fill      = _fill(bal_bg)
    status_cell.font      = _font(bold=True, size=11, color=bal_fg)
    status_cell.alignment = _align(h="center")
    status_cell.border    = _border_thin()
    ws.row_dimensions[r].height = 24
    r += 2

    # ── Summary stats ──────────────────────────────────────────
    _section_header(ws, "  QUICK SUMMARY", NCOLS, r); r += 1
    summary_rows = [
        ("Period",           period,                              None),
        ("Total Collected",  f"Rs. {total_collected:,.0f}",       C_PAID_FG),
        ("Total Expenses",   f"Rs. {total_expenses:,.0f}",        C_DEFICIT_FG),
        ("Net Balance",      f"Rs. {net_balance:,.0f} ({bal_label})", bal_fg),
        ("Expense Categories", str(len(breakdown)),               None),
    ]
    for i, (label, value, fg) in enumerate(summary_rows):
        bg = C_ALT_ROW if i % 2 == 0 else None
        _data_cell(ws, r, 1, "", bg=bg)
        _data_cell(ws, r, 2, label, bg=bg, bold=True)
        _data_cell(ws, r, 3, value, bg=bg, bold=False,
                   fg=fg or "000000", h_align="right")
        ws.cell(row=r, column=4).fill = _fill(bg) if bg else PatternFill()
        r += 1

    r += 1
    _generated_note(ws, r, NCOLS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
